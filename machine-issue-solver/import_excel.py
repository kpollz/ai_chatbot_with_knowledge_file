#!/usr/bin/env python3
"""
Import Excel data into Machine Issue Solver database via API.

Excel columns (in order):
  STT, Line, Team, Machine, Location, Serial, Date, Start Time, Stop Time,
  Total Time, Week, Year, Hiện tượng, Nguyên nhân, Khắc phục, PIC, User Input

Usage:
  python import_excel.py <excel_file> [--api-url http://localhost:8888] [--sheet 0] [--dry-run] [--start-row 2]

Examples:
  # Dry run (no data written, just print what would be imported)
  python import_excel.py data.xlsx --dry-run

  # Import starting from row 5 (skip first 4 rows)
  python import_excel.py data.xlsx --start-row 5

  # Import with custom API URL
  python import_excel.py data.xlsx --api-url http://192.168.1.100:8888
"""

import argparse
import sys
import time
from pathlib import Path

import httpx
import openpyxl


# Excel column index → API field mapping (0-based)
COLUMN_MAP = {
    0: None,             # STT — skip
    1: "LineName",       # Line
    2: "TeamName",       # Team
    3: "MachineName",    # Machine
    4: "Location",       # Location
    5: "Serial",         # Serial
    6: "Date",           # Date
    7: "start_time",     # Start Time
    8: "stop_time",      # Stop Time
    9: "total_time",     # Total Time
    10: "Week",          # Week (int)
    11: "Year",          # Year (int)
    12: "hien_tuong",    # Hiện tượng
    13: "nguyen_nhan",   # Nguyên nhân
    14: "khac_phuc",     # Khắc phục
    15: "PIC",           # PIC
    16: "user_input",    # User Input
}

# Fields that should be converted to int
INT_FIELDS = {"Week", "Year"}

# Fields that should be formatted as ISO date
DATE_FIELDS = {"Date"}

# Fields that are required (must have a non-empty value)
REQUIRED_FIELDS = ["LineName", "TeamName", "MachineName"]


def parse_row(row) -> dict:
    """Parse an Excel row into an API-compatible dict."""
    data = {}
    for col_idx, api_field in COLUMN_MAP.items():
        if api_field is None:
            continue
        if col_idx >= len(row):
            continue

        cell_value = row[col_idx].value
        if cell_value is None:
            continue

        # Convert to string, strip whitespace
        value = str(cell_value).strip()
        if not value:
            continue

        # Convert numeric fields
        if api_field in INT_FIELDS:
            try:
                value = int(float(value))
            except (ValueError, TypeError):
                print(f"  ⚠ Cannot convert '{value}' to int for field '{api_field}', skipping")
                continue

        # Convert date fields to ISO format (YYYY-MM-DD)
        if api_field in DATE_FIELDS:
            try:
                # Handle datetime/date objects from Excel
                from datetime import datetime, date
                if isinstance(cell_value, datetime):
                    value = cell_value.strftime("%Y-%m-%d")
                elif isinstance(cell_value, date):
                    value = cell_value.strftime("%Y-%m-%d")
                # If already a string, keep as is (API will try multiple formats)
            except Exception as e:
                print(f"  ⚠ Cannot convert date for field '{api_field}': {e}, skipping")
                continue

        data[api_field] = value

    return data


def validate_row(data: dict, row_num: int) -> bool:
    """Check if a row has all required fields."""
    missing = [f for f in REQUIRED_FIELDS if f not in data or not data[f]]
    if missing:
        print(f"  ⚠ Row {row_num}: Missing required fields: {', '.join(missing)} — SKIPPED")
        return False
    return True


def import_excel(excel_path: str, api_url: str, sheet_name_or_idx=0,
                 dry_run=False, start_row=2):
    """
    Read Excel file and import each row via POST /issues/import.

    Args:
        excel_path: Path to .xlsx file
        api_url: Base URL of Issue API
        sheet_name_or_idx: Sheet name (str) or index (int, 0-based)
        dry_run: If True, print rows without calling API
        start_row: 1-based row number to start reading (default 2, skip header)
    """
    print(f"📖 Reading: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Get sheet
    if isinstance(sheet_name_or_idx, int):
        sheet = wb.worksheets[sheet_name_or_idx]
    else:
        sheet = wb[sheet_name_or_idx]

    print(f"   Sheet: {sheet.title}")
    print(f"   API:   {api_url}")
    print(f"   Mode:  {'DRY RUN (no data written)' if dry_run else 'LIVE IMPORT'}")
    print(f"   Start: Row {start_row}")
    print()

    # Stats
    total = 0
    success = 0
    skipped = 0
    errors = 0
    duplicates = 0
    created_lines = 0
    created_teams = 0
    created_machines = 0

    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
        total += 1
        data = parse_row(row)

        if not data:
            print(f"  Row {row_idx}: Empty — SKIPPED")
            skipped += 1
            continue

        if not validate_row(data, row_idx):
            skipped += 1
            continue

        if dry_run:
            print(f"  Row {row_idx}: {data.get('LineName')} / {data.get('TeamName')} / "
                  f"{data.get('MachineName')} — {data.get('hien_tuong', '')[:50]}")
            success += 1
            continue

    # Call API - create client once outside loop
    try:
        with httpx.Client(timeout=30) as client:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
                total += 1
                data = parse_row(row)

                if not data:
                    print(f"  Row {row_idx}: Empty — SKIPPED")
                    skipped += 1
                    continue

                if not validate_row(data, row_idx):
                    skipped += 1
                    continue

                if dry_run:
                    print(f"  Row {row_idx}: {data.get('LineName')} / {data.get('TeamName')} / "
                          f"{data.get('MachineName')} — {data.get('hien_tuong', '')[:50]}")
                    success += 1
                    continue

                # Call API
                try:
                    response = client.post(f"{api_url}/issues/import", json=data)

                    if response.status_code == 201:
                        result = response.json()
                        
                        # Check if duplicate
                        if result.get("is_duplicate"):
                            print(f"  ⚠️ Row {row_idx}: DUPLICATE Issue #{result['IssueID']} (skipped)")
                            duplicates += 1
                            continue
                        
                        flags = []
                        if result.get("created_line"):
                            created_lines += 1
                            flags.append("NEW LINE")
                        if result.get("created_team"):
                            created_teams += 1
                            flags.append("NEW TEAM")
                        if result.get("created_machine"):
                            created_machines += 1
                            flags.append("NEW MACHINE")

                        flag_str = f" [{', '.join(flags)}]" if flags else ""
                        print(f"  ✅ Row {row_idx}: Issue #{result['IssueID']}{flag_str}")
                        success += 1
                    else:
                        print(f"  ❌ Row {row_idx}: HTTP {response.status_code} — {response.text[:200]}")
                        errors += 1

                except Exception as e:
                    print(f"  ❌ Row {row_idx}: {e}")
                    errors += 1

    except httpx.ConnectError:
        print(f"\n🚫 Cannot connect to API at {api_url}. Is the service running?")

    wb.close()

    # Summary
    print()
    print("=" * 50)
    print(f"📊 Summary:")
    print(f"   Total rows:   {total}")
    print(f"   Successful:   {success}")
    print(f"   Duplicates:   {duplicates}")
    print(f"   Skipped:      {skipped}")
    print(f"   Errors:       {errors}")
    if not dry_run:
        print(f"   New Lines:    {created_lines}")
        print(f"   New Teams:    {created_teams}")
        print(f"   New Machines: {created_machines}")
    print("=" * 50)


def main():
    parser = argparse.ArgumentParser(
        description="Import Excel data into Machine Issue Solver"
    )
    parser.add_argument(
        "excel_file",
        help="Path to Excel file (.xlsx)"
    )
    parser.add_argument(
        "--api-url",
        default="http://localhost:8888",
        help="Issue API base URL (default: http://localhost:8888)"
    )
    parser.add_argument(
        "--sheet",
        default="0",
        help="Sheet name or 0-based index (default: 0 = first sheet)"
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Row number to start reading, 1-based (default: 2, skip header)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print rows without importing (for testing)"
    )

    args = parser.parse_args()

    # Check file exists
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)

    # Parse sheet argument (int = index, str = name)
    try:
        sheet = int(args.sheet)
    except ValueError:
        sheet = args.sheet

    import_excel(
        excel_path=str(excel_path),
        api_url=args.api_url,
        sheet_name_or_idx=sheet,
        dry_run=args.dry_run,
        start_row=args.start_row,
    )


if __name__ == "__main__":
    main()